from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from vat_reconciliation_engine.config_loader import load_config
from vat_reconciliation_engine.exporter import build_exception_workbook_bytes
from vat_reconciliation_engine.ingestion import load_sample_data
from vat_reconciliation_engine.reconciliation import build_reconciliation


def test_build_exception_workbook_bytes_contains_expected_sheets(project_root):
    config = load_config(project_root / "config" / "vat_generic.yml")
    data = load_sample_data(project_root / "data" / "sample", config)
    result = build_reconciliation(data, config)

    workbook_bytes = build_exception_workbook_bytes(data, result)
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    assert workbook.sheetnames == [
        "summary",
        "output_vat",
        "input_vat",
        "gl_movement",
        "payments_refunds",
        "exceptions",
        "reconciling_items",
    ]
